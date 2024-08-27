import streamlit as st
import pandas as pd
import numpy as np
import re
import logging
from scipy import stats
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import levene, f_oneway, kruskal, median_test
from statsmodels.stats.weightstats import CompareMeans, DescrStatsW
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from io import BytesIO
import io
from PIL import Image
from openpyxl.drawing.image import Image as OpenpyxlImage

# Initialize logging
logging.basicConfig(
    filename='user_logs.log',
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
)

# Initialize session state variables for tracking logging
if 'logged_access' not in st.session_state:
    st.session_state.logged_access = False

if 'logged_upload' not in st.session_state:
    st.session_state.logged_upload = False

if 'logged_download' not in st.session_state:
    st.session_state.logged_download = False

# Function to log user actions
def log_action(action_message):
    """
    Logs user actions to a specified log file.

    Parameters:
    action_message (str): The message detailing the user's action.
    """
    logging.info(f"User action: {action_message}")

# Log access to the app
if not st.session_state.logged_access:
    log_action("Accessed App")
    st.session_state.logged_access = True

# CSS Code for changing color to blue for dropdowns
st.markdown("""
    <style>
    /* Change the background color of selected options */
    div[data-baseweb="select"] span {
        background-color: #12239E !important;
        color: white !important;
    }

    /* Change the color of the text within the selected options */
    div[data-baseweb="select"] span > div {
        color: white !important;
    }
    </style>
    """, unsafe_allow_html=True)

# Set global pandas option for float formatting to 3 decimal places
pd.options.display.float_format = '{:.3f}'.format

def convert_columns_to_numeric(df, exceptions=[]):
    """
    Converts columns in a DataFrame to numeric types, excluding specified exceptions.

    Parameters:
    df (pd.DataFrame): The input DataFrame.
    exceptions (list): List of column names to exclude from conversion.

    Returns:
    pd.DataFrame: The modified DataFrame with columns converted to numeric types.
    """
    for col in df.columns:
        if col not in exceptions:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df

def filter_dataframe(selected_values):
    """
    Filters the DataFrame based on selected values from the 'DESCRIPTION' column.

    Parameters:
    selected_values (list): The list of selected 'DESCRIPTION' values.

    Returns:
    pd.DataFrame: The filtered DataFrame.
    """
    if 'updated_data' in st.session_state:
        updated_data = st.session_state.updated_data
        filtered_df = updated_data[updated_data['DESCRIPTION'].isin(selected_values)]
        st.write(filtered_df)
        return filtered_df
    else:
        st.error("No data available. Please upload the file again.")
        return pd.DataFrame()  # Return an empty DataFrame if no data is available

def select_columns(filtered_df, selected_columns):
    """
    Selects specified columns from a filtered DataFrame.

    Parameters:
    filtered_df (pd.DataFrame): The filtered DataFrame.
    selected_columns (list): The list of columns to select.

    Returns:
    pd.DataFrame: The final DataFrame containing only the selected columns.
    """
    final_df = filtered_df[list(selected_columns)]
    st.write(final_df)
    return final_df

def create_column_dropdown(filtered_df):
    """
    Creates a dropdown list of columns from a filtered DataFrame.

    Parameters:
    filtered_df (pd.DataFrame): The filtered DataFrame.

    Returns:
    list: A list of column names.
    """
    return filtered_df.columns.tolist()

def set_target_column(final_df, selected_column):
    """
    Sets a target column in the DataFrame and renames it to 'GROUP'.

    Parameters:
    final_df (pd.DataFrame): The input DataFrame.
    selected_column (str): The name of the column to set as the target.

    Returns:
    pd.DataFrame: The modified DataFrame with the target column renamed to 'GROUP'.
    """
    global final_df_global
    df_modified = final_df.copy()
    df_modified.rename(columns={selected_column: 'GROUP'}, inplace=True)
    cols = list(df_modified.columns)
    cols.append(cols.pop(cols.index('GROUP')))
    df_modified = df_modified[cols]
    columns_to_exclude = ['SAP-Desc', 'DESCRIPTION', 'GROUP', 'Attributes']
    for col in df_modified.columns:
        if col not in columns_to_exclude:
            df_modified[col] = pd.to_numeric(df_modified[col], errors='coerce')
    final_df_global = df_modified
    st.write(final_df_global)
    return final_df_global

def extract_attributes(description, granularity_choice):
    """
    Extracts attributes from the description based on the selected granularity.

    Parameters:
    description (str): The description string.
    granularity_choice (str): The chosen granularity level.

    Returns:
    str: The extracted attributes.
    """
    if granularity_choice == 'Upto Dimensions':
        parts = description.split(" ")
        attributes = " ".join(parts[:-2]) + "-" + parts[-2] + " " + parts[-1]
        return attributes
    else:
        match = re.match(r"(\d+\.\d+|\d+|\b[a-zA-Z]\b)\s+(.+?)\s+\d", description)
        if match:
            thickness = match.group(1)
            material_type = match.group(2)
            return f"{thickness} {material_type}"
        return None

def calculate_statistics(df, description_column, group_column, granularity_choice):
    """
    Calculates statistics for the DataFrame grouped by 'Attributes' and 'GROUP'.

    Parameters:
    df (pd.DataFrame): The input DataFrame.
    description_column (str): The column containing descriptions.
    group_column (str): The column containing group labels.
    granularity_choice (str): The selected granularity level.

    Returns:
    pd.DataFrame: A DataFrame containing the calculated statistics.
    """
    df['Attributes'] = df[description_column].apply(lambda x: extract_attributes(x, granularity_choice))
    df = df.dropna(subset=['Attributes'])
    grouped = df.groupby(['Attributes', group_column])
    
    def group_stats(group):
        group = group.drop(columns=[description_column, 'Attributes', group_column])
        stats = {
            'N': group.shape[0],
            'N*': group.isnull().sum(),
            'Mean': group.mean(numeric_only=True),
            'StDev': group.std(numeric_only=True),
            'Minimum': group.min(numeric_only=True),
            'Median': group.median(numeric_only=True),
            'Maximum': group.max(numeric_only=True),
            'Skewness': group.skew(numeric_only=True)
        }
        return pd.DataFrame(stats)
    
    stats_df = grouped.apply(group_stats).reset_index()
    return stats_df

def calculate_and_display_statistics(final_df, granularity_choice):
    """
    Calculates and displays statistics for the final DataFrame.

    Parameters:
    final_df (pd.DataFrame): The input DataFrame.
    granularity_choice (str): The selected granularity level.
    """
    global table1_df_global
    if granularity_choice == 'Upto Top and Bottom Facer':
        description_column = 'DESCRIPTION'
        group_column = 'GROUP'
    else:
        description_column = 'SAP-Desc'
        group_column = 'GROUP'
    
    stats_df = calculate_statistics(final_df, description_column, group_column, granularity_choice)
    stats_df.rename(columns={'level_2': 'Variable'}, inplace=True)
    columns_order = ['Attributes', 'GROUP', 'Variable', 'N', 'N*', 'Mean', 'StDev', 'Minimum', 'Median', 'Maximum', 'Skewness']
    stats_df = stats_df[columns_order]
    stats_df = stats_df.round(3)
    pd.options.display.float_format = '{:.6f}'.format
    for col in ['Mean', 'StDev', 'Minimum', 'Median', 'Maximum', 'Skewness']:
        if col in stats_df.columns:
            stats_df[col] = stats_df[col].apply(lambda x: f"{x:.3f}" if pd.notnull(x) else x)
    
    table1_df = stats_df.sort_values(by=['Attributes', 'Variable', 'GROUP'])
    table1_df = table1_df.reset_index().drop(columns=['index'])
    table1_df_global = table1_df
    st.write("**Statistics DataFrame:**")
    st.write(table1_df)

def grubbs_test(data, alpha=0.05):
    """
    Performs Grubbs' test for outliers in the data.

    Parameters:
    data (array-like): The input data.
    alpha (float): The significance level for the test.

    Returns:
    tuple: Grubbs' test statistic and the critical value.
    """
    n = len(data)
    if n < 3:
        return None, None
    mean_y = np.mean(data)
    std_y = np.std(data, ddof=1)
    if std_y == 0:
        return None, None
    numerator = np.max(np.abs(data - mean_y))
    grubbs_statistic = numerator / std_y
    t_dist = stats.t.ppf(1 - alpha / (2 * n), n - 2)
    critical_value = ((n - 1) / np.sqrt(n)) * np.sqrt(t_dist ** 2 / (n - 2 + t_dist ** 2))
    return grubbs_statistic, critical_value

def detect_outliers(final_df, granularity_choice, alpha=0.05):
    """
    Detects outliers in the DataFrame using Grubbs' test.

    Parameters:
    final_df (pd.DataFrame): The input DataFrame.
    granularity_choice (str): The selected granularity level.
    alpha (float): The significance level for the test.

    Returns:
    pd.DataFrame: A DataFrame containing detected outliers.
    """
    global outliers_df_global
    if granularity_choice == 'Upto Top and Bottom Facer':
        column_to_use = 'DESCRIPTION'
    else:
        column_to_use = 'SAP-Desc'
    
    final_df = final_df.copy()
    final_df['Attributes'] = final_df[column_to_use].apply(lambda x: extract_attributes(x, granularity_choice))
    outliers_list = []
    grouped = final_df.groupby('Attributes')
    
    for attr, group_df in grouped:
        for column in group_df.columns:
            if group_df[column].dtype != object and column != 'GROUP':
                data_column = pd.to_numeric(group_df[column], errors='coerce').dropna().values
                while True:
                    grubbs_statistic, critical_value = grubbs_test(data_column, alpha)
                    if grubbs_statistic is None or critical_value is None:
                        break
                    if grubbs_statistic > critical_value:
                        outlier = data_column[np.argmax(np.abs(data_column - np.mean(data_column)))]
                        outliers_list.append({'Attributes': attr, 'Column': column, 'Outlier': outlier})
                        data_column = data_column[data_column != outlier]
                    else:
                        break
    
    outliers_df_global = pd.DataFrame(outliers_list)
    st.write("**Detected Outliers DataFrame:**")
    st.write(outliers_df_global)

def add_attributes_column(final_df, granularity_choice):
    """
    Adds an 'Attributes' column to the DataFrame based on the selected granularity.

    Parameters:
    final_df (pd.DataFrame): The input DataFrame.
    granularity_choice (str): The selected granularity level.

    Returns:
    pd.DataFrame: The modified DataFrame with the 'Attributes' column added.
    """
    if granularity_choice == 'Upto Top and Bottom Facer':
        final_df['Attributes'] = final_df['DESCRIPTION'].apply(lambda x: extract_attributes(x, granularity_choice))
    else:
        final_df['Attributes'] = final_df['SAP-Desc'].apply(lambda x: extract_attributes(x, granularity_choice))
    return final_df

def prompt_outlier_removal(final_df, outliers_df, granularity_choice):
    """
    Prompts the user to remove detected outliers from the DataFrame.

    Parameters:
    final_df (pd.DataFrame): The input DataFrame.
    outliers_df (pd.DataFrame): The DataFrame containing detected outliers.
    granularity_choice (str): The selected granularity level.

    Returns:
    pd.DataFrame: The modified DataFrame after outlier removal.
    """
    final_df = add_attributes_column(final_df, granularity_choice)
    st.write(outliers_df)
    
    if outliers_df.empty:
        st.write("No outliers detected.")
        return final_df
    
    for index, row in outliers_df.iterrows():
        attr = row['Attributes']
        column = row['Column']
        outlier_value = row['Outlier']
        
        decision = st.radio(f"Do you want to remove the outlier {outlier_value} from {attr} in column {column}?", ('Yes', 'No'), key=index)
        
        if decision == 'Yes':
            final_df = final_df[~((final_df[column] == outlier_value) & (final_df['Attributes'] == attr))]
            st.write(f"Outlier {outlier_value} from {attr} in column {column} removed.")
        else:
            st.write(f"Outlier {outlier_value} from {attr} in column {column} not removed.")
    
    return final_df

def interactive_sap_filter_and_select():
    """
    Provides an interactive interface for filtering and selecting SAPs and merging descriptions.

    Returns:
    dict: A dictionary containing the filtered DataFrame, final DataFrame, and merged SAP descriptions.
    """
    if 'updated_data' in st.session_state:
        updated_data = st.session_state.updated_data
        result = {'filtered_df': None, 'final_df': None, 'merged_sap_desc': None}
        sap_options = updated_data['SAP'].unique().tolist()
        sap_options.insert(0, 'Select SAPs')

        selected_saps = st.multiselect('Select SAPs', sap_options)
        if 'Select SAPs' not in selected_saps and selected_saps:
            filtered_df = updated_data[updated_data['SAP'].isin(selected_saps)]
            st.write(filtered_df)
            
            merge_count = st.number_input('Number of Merges', min_value=1, step=1)
            for merge_num in range(1, merge_count + 1):
                merge_selected_saps = st.multiselect(f'Select SAPs for Merge {merge_num}', selected_saps, key=f"merge_{merge_num}")
                if 'Select SAPs' not in merge_selected_saps and len(merge_selected_saps) >= 2:
                    merged_description = " / ".join(filtered_df[filtered_df['SAP'].isin(merge_selected_saps)]['DESCRIPTION'].unique())
                    merge_key = f"Merge_{merge_num}_-" + "-".join(map(str, merge_selected_saps))
                    updated_data.loc[updated_data['SAP'].isin(merge_selected_saps), 'SAP-Desc'] = merged_description
                    updated_data.loc[updated_data['SAP'].isin(merge_selected_saps), 'Merge-Key'] = merge_key
                    st.write(f"Merged Description: {merged_description}")
            
            filtered_df['SAP-Desc'] = filtered_df['SAP'].map(updated_data.drop_duplicates(subset='SAP').set_index('SAP')['SAP-Desc'])
            filtered_df['Merge-Key'] = filtered_df['SAP'].map(updated_data.drop_duplicates(subset='SAP').set_index('SAP')['Merge-Key'])
            column_options = create_column_dropdown(filtered_df)
            column_options.insert(0, 'Select Columns')
            selected_columns = st.multiselect('Select Columns', column_options)
            if 'Select Columns' not in selected_columns and selected_columns:
                final_df = select_columns(filtered_df, selected_columns)
                target_column = st.selectbox('Select Target Column', ['Select Target Column'] + final_df.columns.tolist(), index=0)
                if target_column != 'Select Target Column':
                    new_df = set_target_column(final_df, target_column)
                    calculate_and_display_statistics(new_df, granularity_choice)
                    detect_outliers(new_df, granularity_choice)
    else:
        st.error("No data available. Please upload the file again.")

def anderson_darling_test(data):
    """
    Performs the Anderson-Darling test for normality on the data.

    Parameters:
    data (array-like): The input data.

    Returns:
    tuple: The Anderson-Darling statistic, critical values, and significance level.
    """
    data = np.asarray(data).flatten()
    result = stats.anderson(data, dist='norm')
    return result.statistic, result.critical_values, result.significance_level

def find_representative_description(descriptions, threshold=80):
    """
    Finds a representative description from a list of descriptions using fuzzy matching.

    Parameters:
    descriptions (list): The list of descriptions.
    threshold (int): The similarity threshold for matching.

    Returns:
    str: The representative description.
    """
    unique_descriptions = np.unique(descriptions)
    representative_description = unique_descriptions[0]
    for desc in unique_descriptions[1:]:
        if fuzz.ratio(representative_description, desc) >= threshold:
            continue
        else:
            representative_description = desc
            break
    return representative_description

def plot_standard_deviations(column, group_labels, means, std_devs, p_value_1, p_value_2):
    """
    Plots standard deviations with multiple comparison intervals.

    Parameters:
    column (str): The column name.
    group_labels (list): The list of group labels.
    means (list): The list of means.
    std_devs (list): The list of standard deviations.
    p_value_1 (float): The p-value from Levene's test.
    p_value_2 (float): The p-value from ANOVA or Welch's test.

    Returns:
    plt.Figure: The matplotlib figure object.
    """
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.errorbar(means, group_labels, xerr=std_devs, fmt='o', capsize=5, capthick=2, elinewidth=2, linestyle='None')
    ax.set_title(f'Test for Equal Variances: {column} vs GROUP\nMultiple comparison intervals for the standard deviation, α = 0.05')
    ax.set_ylabel('GROUP')
    ax.set_xlabel('Standard Deviation')
    plt.figtext(0.95, 0.9, f"Multiple Comparisons\nP-Value (Levene's Test)\nP-Value: {p_value_1:.3f}", bbox={"facecolor": "lightgray", "alpha": 0.5, "pad": 5})
    plt.figtext(0.95, 0.8, f"P-Value (ANOVA or Welch)\nP-Value: {p_value_2:.3f}", bbox={"facecolor": "lightgray", "alpha": 0.5, "pad": 5})
    ax.set_yticks(range(len(group_labels)))
    ax.set_yticklabels(group_labels)
    plt.close(fig)
    return fig

def handle_choice_1(choice):
    """
    Handles the user's choice of test after skewness checking.

    Parameters:
    choice (str): The selected test option.
    """
    global user_choice_1
    user_choice_1 = choice
    if user_choice_1 == 'Equal Variance Test':
        equal_variance_test()
    elif user_choice_1 == 'Non-Parametric Test':
        check_extreme_outliers()

def handle_choice_2(attr, group_df):
    """
    Handles the user's choice of statistical test for a specific attribute.

    Parameters:
    attr (str): The attribute value.
    group_df (pd.DataFrame): The DataFrame for the specific attribute group.

    Returns:
    list: A list of results from the chosen statistical test.
    """
    def handle_choice_test(choice):
        global user_choice_2, global_results
        user_choice_2 = choice
        st.write(f"User choice for {attr}: {user_choice_2}")

        global_results = [result for result in global_results if not (result['Attributes'] == attr and result['Test Conducted'] == user_choice_2)]
        
        if user_choice_2 == "Mood's Median":
            results = moods_median_test(group_df, attr)
        elif user_choice_2 == 'Kruskal-Wallis':
            results = kruskal_wallis_test(group_df, attr)
        
        global_results.extend(results)
        results_df = pd.DataFrame(results)
        st.write(results_df)
        return results

    choice = st.selectbox("Choose Test:", ['Select', "Mood's Median", 'Kruskal-Wallis'], key=f"test_{attr}")
    if choice != 'Select':
        handle_choice_test(choice)

global_results = []

def check_skewness_and_proceed():
    """
    Checks skewness and guides the user to choose the appropriate test based on the result.
    """
    table1_df_global['Skewness'] = pd.to_numeric(table1_df_global['Skewness'], errors='coerce')
    skewness_check = table1_df_global['Skewness'].apply(lambda x: -2 <= x <= 2 if pd.notna(x) else True)
    if skewness_check.all():
        st.write("Skewness for all records is between -2 and 2. Proceed to user choice.")
        handle_choice_1({"Select": "Select", "Parametric Test (ANOVA)": "Equal Variance Test", "Non-Parametric Test": "Non-Parametric Test"}[st.selectbox('Choose Test:', ['Select', 'Parametric Test (ANOVA)', 'Non-Parametric Test'], key='skewness_test')])
    else:
        st.write("Skewness for some records is outside the range of -2 and 2. Suggesting Non-Parametric Test (NPT) instead.")
        handle_choice_1({"Select": "Select", "Parametric Test (ANOVA)": "Equal Variance Test", "Non-Parametric Test": "Non-Parametric Test"}[st.selectbox('Choose Test:', ['Select', 'Parametric Test (ANOVA)', 'Non-Parametric Test'], key='skewness_test')])

def perform_anderson_darling_test(filtered_data):
    """
    Performs the Anderson-Darling test for normality and guides the user to the appropriate test.

    Parameters:
    filtered_data (pd.DataFrame): The filtered DataFrame.
    """
    numerical_columns = [col for col in filtered_data.columns if pd.api.types.is_numeric_dtype(filtered_data[col])]
    data_for_testing = filtered_data[numerical_columns].values.flatten()
    data_for_testing = data_for_testing[~np.isnan(data_for_testing)]
    data_for_testing = data_for_testing[np.isfinite(data_for_testing)]
    statistic, critical_values, significance_level = anderson_darling_test(data_for_testing)
    p_value_threshold = 0.05
    if statistic < critical_values[2]:
        result = "Yes"        
        st.markdown(f"<span style='color:#12239E; font-weight:bold;'>Anderson Normality Test Result = {result} (p>0.05)</span>", unsafe_allow_html=True)
        st.markdown("<span style='color:#12239E; font-weight:bold;'>Valid Null Hypothesis: Data is Normally Distributed</span>", unsafe_allow_html=True)
        check_skewness_and_proceed()
    else:
        result = "No"
        st.markdown(f"<span style='color:#12239E; font-weight:bold;'>Anderson Normality Test Result = {result} (p<=0.05)</span>", unsafe_allow_html=True)
        st.markdown("<span style='color:#12239E; font-weight:bold;'>Invalid Null Hypothesis: Data is not Normally Distributed</span>", unsafe_allow_html=True)
        perform_mgrt_test()

def perform_mgrt_test():
    """
    Performs the Minimum Group Records Test (MGRT) and guides the user to the next step based on the results.
    """
    group_counts = final_df_global.groupby(['Attributes', 'GROUP']).size().reset_index(name='Number of Data Points')
    st.write("Group counts:\n", group_counts)
    result_MGRT = "Yes"
    
    for attr, group_count in group_counts.groupby('Attributes'):  # Use 'Attributes' instead of level=0
        attribute_value = attr  # attr now contains the actual 'Attributes' value
        if 2 <= len(group_count) <= 9:
            st.write(f"Number of unique groups for {attribute_value} is between 2 and 9.")
            if not all(group_count['Number of Data Points'] > 15):
                result_MGRT = "No"
                st.write(f"At least one group in {attribute_value} has 15 or fewer records.")
            else:
                st.write(f"All groups in {attribute_value} have more than 15 records.")
        elif 10 <= len(group_count) <= 12:
            st.write(f"Number of unique groups for {attribute_value} is between 10 and 12.")
            if not all(group_count['Number of Data Points'] > 20):
                result_MGRT = "No"
                st.write(f"At least one group in {attribute_value} has 20 or fewer records.")
            else:
                st.write(f"All groups in {attribute_value} have more than 20 records.")
        else:
            st.write(f"Number of unique groups for {attribute_value} is outside the specified range (2-12).")
    
    st.markdown(f"<span style='color:#12239E; font-weight:bold;'>Result from Minimum Group Records Test = {result_MGRT}</span>", unsafe_allow_html=True)
    
    if result_MGRT == "Yes":
        check_skewness_and_proceed()
    else:
        handle_choice_1({"Select": "Select", "Parametric Test (ANOVA)": "Equal Variance Test", "Non-Parametric Test": "Non-Parametric Test"}[st.selectbox('Choose Test:', ['Select', 'Parametric Test (ANOVA)', 'Non-Parametric Test'], key='mgr_test')])

def equal_variance_test():
    """
    Conducts the equal variance test using Levene's test and ANOVA or Welch's test, and displays the results.
    """
    global global_levene_results, global_anova_welch_results
    descriptions = final_df_global['Attributes']
    representative_description = find_representative_description(descriptions)
    levene_results = []
    anova_welch_results = []
    numeric_columns = final_df_global.select_dtypes(include='number').columns
    grouped = final_df_global.groupby('Attributes')
    
    for attr, group_df in grouped:
        for column in numeric_columns:
            if column != 'GROUP':
                groups = group_df.groupby('GROUP')[column].apply(lambda x: x.dropna().values)
                group_labels = [label for label, group in zip(groups.index, groups) if len(group) > 0]
                means = [np.mean(group) for group in groups if len(group) > 0]
                std_devs = [np.std(group) for group in groups if len(group) > 0]
                if not means or not std_devs:
                    continue
                try:
                    stat, p_value_1 = levene(*[group for group in groups if len(group) > 0])
                    if np.isnan(p_value_1):
                        raise ValueError("NaN P-Value")
                except:
                    p_value_1 = np.nan
                    levene_results.append({
                        'Column': column,
                        'Levene P-Value': 'N/A',
                        'Result': 'N/A',
                        'Attributes': attr
                    })
                    anova_welch_results.append({
                        'Column': column,
                        'ANOVA/Welch P-Value': 'N/A',
                        'Result': 'N/A',
                        'Test Type': 'N/A',
                        'Attributes': attr
                    })
                    continue
                levene_result = 'Equal' if p_value_1 > 0.05 else 'Diff'
                levene_results.append({
                    'Column': column,
                    'Levene P-Value': p_value_1,
                    'Levene Statistic': stat,
                    'Result': levene_result,
                    'Attributes': attr
                })
                try:
                    if p_value_1 > 0.05:
                        anova_result = f_oneway(*[group for group in groups if len(group) > 0])
                        p_value_2 = anova_result.pvalue
                        if np.isnan(p_value_2):
                            raise ValueError("NaN P-Value")
                        if p_value_2 > 0.05:
                            result = 'Equal'
                            test_type = 'ANOVA'
                        else:
                            result = 'Diff'
                            test_type = 'ANOVA'
                    else:
                        desc_stats = [DescrStatsW(group) for group in groups if len(group) > 0]
                        cm = CompareMeans(*desc_stats)
                        welch_result = cm.ttest_ind(usevar='unequal')
                        p_value_2 = welch_result[1]
                        if np.isnan(p_value_2):
                            raise ValueError("NaN P-Value")
                        if p_value_2 > 0.05:
                            result = 'Equal'
                            test_type = 'Welch ANOVA'
                        else:
                            result = 'Diff'
                            test_type = 'Welch ANOVA'
                except:
                    p_value_2 = np.nan
                    result = 'N/A'
                    test_type = 'N/A'
                fig = plot_standard_deviations(column, group_labels, means, std_devs, p_value_1, p_value_2)
                plots_dict.setdefault('standard_deviations', []).append((f'{attr}_{column}', fig))
                anova_welch_results.append({
                    'Column': column,
                    'ANOVA/Welch P-Value': p_value_2,
                    'Result': result,
                    'Test Type': test_type,
                    'Attributes': attr
                })
    
    global_levene_results = pd.DataFrame(levene_results)
    global_anova_welch_results = pd.DataFrame(anova_welch_results)
    levene_df_pivot = pd.pivot_table(global_levene_results, index='Attributes', columns='Column', values='Levene P-Value', aggfunc='first')
    anova_welch_df_pivot = pd.pivot_table(global_anova_welch_results, index='Attributes', columns='Column', values='ANOVA/Welch P-Value', aggfunc='first')
    levene_df_simple = global_levene_results.pivot(index='Attributes', columns='Column', values='Result')
    anova_welch_df_simple = global_anova_welch_results.pivot(index='Attributes', columns='Column', values='Result')
    
    st.write("Levene's Test Results (with P-Values): (Equal Variance Test)")
    st.write(round_numeric_columns(levene_df_pivot))
    st.write("\nANOVA/Welch Test Results (with P-Values):")
    st.write(round_numeric_columns(anova_welch_df_pivot))
    st.write("\nLevene's Test Results (Simplified): (Equal Variance Test)")
    st.write(round_numeric_columns(levene_df_simple))
    st.write("\nANOVA/Welch Test Results (Simplified):")
    st.write(round_numeric_columns(anova_welch_df_simple))

def moods_median_test(group_df, attr):
    """
    Conducts Mood's Median test for a specific attribute and returns the results.

    Parameters:
    group_df (pd.DataFrame): The DataFrame for the specific attribute group.
    attr (str): The attribute value.

    Returns:
    list: A list of results from Mood's Median test.
    """
    numeric_columns = group_df.select_dtypes(include='number').columns
    results = []
    
    for column in numeric_columns:
        if column != 'GROUP':
            groups = [group[column].dropna().values for name, group in group_df.groupby('GROUP')]
            groups = [group for group in groups if len(group) > 0]
            if len(groups) < 2:
                continue
            try:
                stat, p_value, med, tbl = median_test(*groups)
                if p_value > 0.05:
                    result = 'Equal'
                else:
                    result = 'Diff'
                
                group_stats = group_df.groupby('GROUP')[column].describe().unstack()
                result_data = {
                    'Attributes': attr,
                    'Column': column,
                    'Group': group_df['GROUP'].unique().tolist(),
                    'Median': group_stats['50%'].to_dict(),
                    'N': group_stats['count'].to_dict(),
                    'P-Value': p_value,
                    'Result': result,
                    'Test Conducted': "Mood's Median"
                }
                results.append(result_data)
            except ValueError as e:
                result_data = {
                    'Attributes': attr,
                    'Column': column,
                    'Group': group_df['GROUP'].unique().tolist(),
                    'Median': np.nan,
                    'N': len(group_df),
                    'P-Value': np.nan,
                    'Result': 'N/A',
                    'Test Conducted': "Mood's Median"
                }
                results.append(result_data)
    return results

def kruskal_wallis_test(group_df, attr):
    """
    Conducts Kruskal-Wallis test for a specific attribute and returns the results.

    Parameters:
    group_df (pd.DataFrame): The DataFrame for the specific attribute group.
    attr (str): The attribute value.

    Returns:
    list: A list of results from Kruskal-Wallis test.
    """
    numeric_columns = group_df.select_dtypes(include='number').columns
    results = []
    
    for column in numeric_columns:
        if column != 'GROUP':
            groups = [group[column].dropna().values for name, group in group_df.groupby('GROUP')]
            groups = [group for group in groups if len(group) > 0]
            if len(groups) < 2:
                continue
            try:
                stat, p_value = kruskal(*groups)
                if p_value > 0.05:
                    result = 'Equal'
                else:
                    result = 'Diff'
                
                group_stats = group_df.groupby('GROUP')[column].describe().unstack()
                result_data = {
                    'Attributes': attr,
                    'Column': column,
                    'Group': group_df['GROUP'].unique().tolist(),
                    'Median': group_stats['50%'].to_dict(),
                    'N': group_stats['count'].to_dict(),
                    'P-Value': p_value,
                    'Result': result,
                    'Test Conducted': "Kruskal-Wallis"
                }
                results.append(result_data)
            except ValueError as e:
                result_data = {
                    'Attributes': attr,
                    'Column': column,
                    'Group': group_df['GROUP'].unique().tolist(),
                    'Median': np.nan,
                    'N': len(group_df),
                    'P-Value': np.nan,
                    'Result': 'N/A',
                    'Test Conducted': "Kruskal-Wallis"
                }
                results.append(result_data)
    return results

def perform_tests_by_attribute():
    """
    Conducts the appropriate statistical tests by attribute and compiles the results.
    """
    all_results = []
    
    for attr in final_df_global['Attributes'].unique():
        group_df = final_df_global[final_df_global['Attributes'] == attr]
        st.write(f"\nChoose the test to conduct for {attr}:")
        results_df = handle_choice_2(attr, group_df)
        if results_df is not None:
            all_results.append(results_df)

    all_results_list = []
    
    for results_df in all_results:
        if isinstance(results_df, list):
            all_results_list.extend(results_df)

    global global_results
    global_results.extend(all_results_list)

def check_extreme_outliers():
    """
    Checks for extreme outliers in the data and guides the user to the next step.

    Returns:
    pd.DataFrame: A DataFrame containing detected extreme outliers.
    """
    def find_extreme_outliers(series):
        q1 = series.quantile(0.25)
        q3 = series.quantile(0.75)
        iqr = q3 - q1
        lower_bound = q1 - 1.5 * iqr
        upper_bound = q3 + 1.5 * iqr
        mean = series.mean()
        outliers = series[(series < lower_bound) | (series > upper_bound)]
        distances = (outliers - mean).abs()
        return outliers, distances

    outlier_present = False
    outliers_list = []
    numeric_columns = final_df_global.select_dtypes(include='number').columns
    grouped = final_df_global.groupby('Attributes')
    
    for attr, group_df in grouped:
        for column in numeric_columns:
            if column != 'GROUP':
                valid_groups = group_df['GROUP'].dropna().unique()
                if len(valid_groups) > 0:
                    groups = group_df.groupby('GROUP')[column].apply(lambda x: x.dropna().values)
                    valid_groups = [group for group in groups if len(group) > 0]
                    if len(valid_groups) < 2:
                        continue
                    fig, ax = plt.subplots(figsize=(10, 6))
                    sns.boxplot(x='GROUP', y=column, data=group_df, ax=ax)
                    plt.title(f"Boxplot of {column} by GROUP for Attributes={attr}")
                    plt.close(fig)
                    plots_dict.setdefault('boxplot', []).append((f'{attr}_{column}', fig))
                
                for name, group in group_df.groupby('GROUP'):
                    outliers, distances = find_extreme_outliers(group[column])
                    if not outliers.empty:
                        outlier_present = True
                        for value, distance in zip(outliers, distances):
                            outliers_list.append({
                                'Column': column,
                                'Group': name,
                                'Attributes': attr,
                                'Outlier': value,
                                'Distance from Mean': distance
                            })
    
    result = "Yes" if outlier_present else "No"
    st.markdown(f"<span style='color:#12239E; font-weight:bold;'>Extreme outliers present: {result}</span>", unsafe_allow_html=True)
    outliers_df = pd.DataFrame(outliers_list)
    st.write(outliers_df)
    perform_tests_by_attribute()

def round_numeric_columns(df, decimals=3):
    """
    Rounds numeric columns in a DataFrame to a specified number of decimal places.

    Parameters:
    df (pd.DataFrame): The input DataFrame.
    decimals (int): The number of decimal places to round to.

    Returns:
    pd.DataFrame: The modified DataFrame with rounded numeric columns.
    """
    return df.applymap(lambda x: round(x, decimals) if isinstance(x, (int, float)) else x)

def display_results(test_choice, trial_group, incumbent_group):
    """
    Displays the results of the statistical analysis and saves them to an Excel file.

    Parameters:
    test_choice (str): The type of test conducted (e.g., 'Non-Parametric Test').
    trial_group (str): The name of the trial group.
    incumbent_group (str): The name of the incumbent group.
    """
    if test_choice == 'Non-Parametric Test':
        npt_results_df = pd.DataFrame(global_results)
        npt_results_df_filtered = npt_results_df.drop_duplicates(subset=['Attributes', 'Column'], keep='last')
        p_values = npt_results_df_filtered[['Attributes', 'Column', 'P-Value']]
    else:
        levene_pivot_pvalue = global_levene_results.pivot(index='Attributes', columns='Column', values='Levene P-Value')
        levene_pivot_statistic = global_levene_results.pivot(index='Attributes', columns='Column', values='Levene Statistic')
        levene_pivot_result = global_levene_results.pivot(index='Attributes', columns='Column', values='Result')

        anova_welch_pivot_pvalue = global_anova_welch_results.pivot(index='Attributes', columns='Column', values='ANOVA/Welch P-Value')
        anova_welch_pivot_result = global_anova_welch_results.pivot(index='Attributes', columns='Column', values='Result')

        equal_variance_choice = "Levene's Test"
        if equal_variance_choice == "Levene's Test":
            p_values = levene_pivot_pvalue.reset_index().melt(id_vars=['Attributes'], var_name='Column', value_name='P-Value').dropna(subset=['P-Value'])
        elif equal_variance_choice == "Welch's ANOVA":
            p_values = anova_welch_pivot_pvalue.reset_index().melt(id_vars=['Attributes'], var_name='Column', value_name='P-Value').dropna(subset=['P-Value'])
        else:
            st.error("Please select a valid test option.")
            return

    result_df = pd.DataFrame()
    trial_data = table1_df_global[table1_df_global['GROUP'] == trial_group]
    incumbent_data = table1_df_global[table1_df_global['GROUP'] == incumbent_group]
    merged_data = pd.merge(trial_data, incumbent_data, on=['Attributes', 'Variable'], suffixes=('_Trial', '_Incumbent'))
    merged_data['Mean_Trial'] = pd.to_numeric(merged_data['Mean_Trial'], errors='coerce')
    merged_data['Mean_Incumbent'] = pd.to_numeric(merged_data['Mean_Incumbent'], errors='coerce')
    merged_data['N_Trial'] = pd.to_numeric(merged_data['N_Trial'], errors='coerce')
    merged_data['N_Incumbent'] = pd.to_numeric(merged_data['N_Incumbent'], errors='coerce')

    result_df['Product'] = merged_data['Attributes']
    result_df['Group'] = merged_data['Variable']
    result_df['Sample Size (Trial)'] = merged_data['N_Trial']
    result_df['Sample Size (Incumbent)'] = merged_data['N_Incumbent']
    result_df['Trial'] = merged_data['Mean_Trial']
    result_df['Incumbent'] = merged_data['Mean_Incumbent']
    result_df['Diff (Trial - Incumbent)'] = merged_data['Mean_Trial'] - merged_data['Mean_Incumbent']

    result_df = result_df.merge(p_values, left_on=['Product', 'Group'], right_on=['Attributes', 'Column'], how='left')
    result_df['P-Value'] = pd.to_numeric(result_df['P-Value'], errors='coerce')
    result_df['Statistically Different'] = result_df['P-Value'] < 0.05
    result_df['Statistically Different'] = result_df['Statistically Different'].apply(lambda x: 'YES' if x else 'NO')
    result_df = result_df.drop(columns=['Column', 'Attributes'])
    
    pivot_df = result_df.melt(id_vars=['Product', 'Group'], 
                   value_vars=['Sample Size (Trial)', 'Sample Size (Incumbent)', 'Trial', 'Incumbent', 'Diff (Trial - Incumbent)', 'P-Value', 'Statistically Different'],
                   var_name='Metric', value_name='Value')

    pivot_df = pivot_df.pivot_table(index=['Product', 'Metric'], columns='Group', values='Value', aggfunc='first').reset_index()
    pivot_df.columns = ['_'.join(col).strip() if type(col) is tuple else col for col in pivot_df.columns.values]
    pivot_df = pivot_df.rename(columns={'Product_': 'Product', 'Metric_': 'Metric'})
    
    metric_order = ['Sample Size (Trial)', 'Sample Size (Incumbent)', 'Trial', 'Incumbent','Diff (Trial - Incumbent)', 'P-Value', 'Statistically Different']
    pivot_df['Metric'] = pd.Categorical(pivot_df['Metric'], categories=metric_order, ordered=True)
    pivot_df = pivot_df.sort_values(by=['Product', 'Metric']).reset_index(drop=True)
    table3_df_global = pivot_df.copy()

    filtered_df = table3_df_global[table3_df_global['Metric'].isin(['P-Value', 'Diff (Trial - Incumbent)'])]
    columns_to_convert = filtered_df.columns.difference(['Product', 'Metric'])
    filtered_df[columns_to_convert] = filtered_df[columns_to_convert].apply(pd.to_numeric, errors='coerce')
    filtered_df = filtered_df.replace(0, np.nan)
    results = []


    for product in filtered_df['Product'].unique():
        product_df = filtered_df[filtered_df['Product'] == product]
        p_values = product_df[product_df['Metric'] == 'P-Value']
        diffs = product_df[product_df['Metric'] == 'Diff (Trial - Incumbent)']
    
        for col in columns_to_convert:
            p_value = p_values[col].values[0] if col in p_values.columns else np.nan
            diff_value = diffs[col].values[0] if col in diffs.columns else np.nan
        
            stat_diff = 'Yes' if not np.isnan(p_value) and p_value < 0.05 else 'No'
        
            assessment = 'Higher' if diff_value > 0 else 'Lower' if diff_value < 0 else 'N/A'
        
            if stat_diff == 'Yes':
                results.append({
                    'Product': product,
                    'Metric': col,
                    'Statistically Diff': stat_diff,
                    'Avg Measured Diff': diff_value,
                    'Assessment': assessment
            })
    table4_df_global = pd.DataFrame(results)

    table2_df_global = []
    if test_choice == 'Non-Parametric Test':
        table2_df_global = npt_results_df_filtered
    else:
        table2_df_global = levene_pivot_result
    
    table2_df_global = table2_df_global.reset_index()
    table2_df_global = table2_df_global.rename(columns={'Attributes':'Product'})
    
    save_results_to_excel(table1_df_global, table2_df_global, table3_df_global, table4_df_global)

def save_results_to_excel(table1_df_global, table2_df_global, table3_df_global, table4_df_global):
    """
    Saves the statistical analysis results to an Excel file with appropriate formatting.

    Parameters:
    table1_df_global (pd.DataFrame): The first table containing statistics.
    table2_df_global (pd.DataFrame): The second table containing test results.
    table3_df_global (pd.DataFrame): The third table containing comparative results.
    table4_df_global (pd.DataFrame): The fourth table containing summary and assessment.
    """
    output_path = BytesIO()
    table1_df_global = round_numeric_columns(table1_df_global)
    table2_df_global = round_numeric_columns(table2_df_global)
    table3_df_global = round_numeric_columns(table3_df_global)
    table4_df_global = round_numeric_columns(table4_df_global)
    
    st.write("Table 1")
    st.write(table1_df_global)
    st.write("Table 2")
    st.write(table2_df_global)
    st.write("Table 3")
    st.write(table3_df_global)
    st.write("Table 4")
    st.write(table4_df_global)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        table1_df_global.to_excel(writer, sheet_name='tables', startrow=0, index=False)

    output_path.seek(0)
    book = load_workbook(output_path)
    sheet = book['tables']

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    def style_range(ws, cell_range, font=None, fill=None, border=None):
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = border

    start_row = len(table1_df_global) + 5

    def write_dataframe_to_excel(sheet, df, start_row, header_font, header_fill, thin_border):
        for col_num, col_name in enumerate(df.columns, start=1):
            cell = sheet.cell(row=start_row, column=col_num, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        style_range(sheet, f'A{start_row}:{chr(64 + len(df.columns))}{start_row}', font=header_font, fill=header_fill, border=thin_border)
        
        start_row += 1

        for row_num, r in enumerate(df.itertuples(index=False), start=start_row):
            for col_num, value in enumerate(r, start=1):
                if isinstance(value, list):
                    value = ', '.join(map(str, value))
                elif isinstance(value, dict):
                    value = ', '.join(f'{k}: {v}' for k, v in value.items())
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        return row_num + 5

    for df in [table2_df_global, table3_df_global, table4_df_global]:
        start_row = write_dataframe_to_excel(sheet, df, start_row, header_font, header_fill, thin_border)

    book.save(output_path)

    def save_plots_to_excel(plots_dict, filename):
        book = load_workbook(filename)
        for plot_type, plots in plots_dict.items():
            if plot_type[:31] in book.sheetnames:
                worksheet = book[plot_type[:31]]
            else:
                worksheet = book.create_sheet(plot_type[:31])
            row = 1
            for value_col, fig in plots:
                if fig is not None:  # Check if the figure is valid
                    img_data = io.BytesIO()
                    fig.savefig(img_data, format='png')
                    img_data.seek(0)
                    image = OpenpyxlImage(img_data)
                    worksheet.add_image(image, f'A{row}')
                    worksheet[f'B{row}'] = value_col
                    row += 50
                else:
                    print(f"Plot for {value_col} in {plot_type} could not be saved because the figure is None.")
        book.save(filename)

    save_plots_to_excel(plots_dict, output_path)

    # Log download action
    if not st.session_state.logged_download:
        log_action("Downloaded Excel file")
        st.session_state.logged_download = True

    st.download_button(
        label="Download Excel File",
        data=output_path.getvalue(),
        file_name="output_tables.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.write("**Output is Processed Successfully**")


# Plotting functions for visualization
def plot_tukey_CIs(data, group_col, value_col):
    """
    Plots Tukey's Simultaneous 95% Confidence Intervals (CIs) for differences of means.

    Parameters:
    data (pd.DataFrame): The input DataFrame.
    group_col (str): The name of the column containing group labels.
    value_col (str): The name of the column containing values to plot.

    Returns:
    plt.Figure: The matplotlib figure object.
    """
    try:
        from statsmodels.stats.multicomp import pairwise_tukeyhsd
        tukey_result = pairwise_tukeyhsd(endog=data[value_col], groups=data[group_col], alpha=0.05)
        tukey_result.plot_simultaneous()
        plt.title(f'Tukey Simultaneous 95% CIs\nDifferences of Means for {value_col}')
        plt.xlabel('Difference of Means')
        plt.grid()
        fig=plt.gcf()
        plt.close(fig)
        return fig
    except Exception as e:
        print(f"Skipping Tukey HSD plot for {value_col}: {e}")
        return None

def plot_interval(data, group_col, value_col):
    """
    Plots the interval plot for the given value column against the group column.

    Parameters:
    data (pd.DataFrame): The input DataFrame.
    group_col (str): The name of the column containing group labels.
    value_col (str): The name of the column containing values to plot.

    Returns:
    plt.Figure: The matplotlib figure object.
    """
    sns.pointplot(x=group_col, y=value_col, data=data, capsize=.1, errwidth=1.5, ci='sd')
    plt.title(f'Interval Plot of {value_col} vs {group_col}\n95% CI for the Mean')
    plt.ylabel(value_col)
    plt.xlabel(group_col)
    plt.grid()
    fig=plt.gcf()
    plt.close(fig)
    return fig
def plot_individual_value(data, group_col, value_col):
    """
    Plots the individual value plot for the given value column against the group column.

    Parameters:
    data (pd.DataFrame): The input DataFrame.
    group_col (str): The name of the column containing group labels.
    value_col (str): The name of the column containing values to plot.

    Returns:
    plt.Figure: The matplotlib figure object.
    """
    if group_col not in data.columns or value_col not in data.columns:
        return None
    if data[group_col].isna().any() or data[value_col].isna().any():
        return None
    try:
        sns.stripplot(x=group_col, y=value_col, data=data, jitter=True)
        sns.pointplot(x=group_col, y=value_col, data=data, join=True, ci='sd', color='red')
        plt.title(f'Individual Value Plot of {value_col} vs {group_col}')
        plt.ylabel(value_col)
        plt.xlabel(group_col)
        plt.grid()
        fig=plt.gcf()
        plt.close(fig)
        return fig
    except Exception as e:
        st.error(f"An error occurred while plotting: {e}")
        return None

def plot_boxplot(data, group_col, value_col):
    """
    Plots the boxplot for the given value column against the group column.

    Parameters:
    data (pd.DataFrame): The input DataFrame.
    group_col (str): The name of the column containing group labels.
    value_col (str): The name of the column containing values to plot.

    Returns:
    plt.Figure: The matplotlib figure object.
    """
    data = data.dropna(subset=[group_col, value_col])
    
    if data[group_col].nunique() < 2:
        print(f"Not enough groups to plot boxplot for {value_col}. Skipping.")
        return None
    
    try:
        sns.boxplot(x=group_col, y=value_col, data=data)
        sns.pointplot(x=group_col, y=value_col, data=data, join=True, ci='sd', color='red')
        plt.title(f'Boxplot of {value_col}')
        plt.ylabel(value_col)
        plt.xlabel(group_col)
        plt.grid()
        fig=plt.gcf()
        plt.close(fig)
        return fig
    except ValueError as e:
        print(f"Error while plotting boxplot for {value_col}: {e}")
        return None

def plot_residuals(data, group_col, value_col):
    """
    Plots the residuals for the given value column against the group column.

    Parameters:
    data (pd.DataFrame): The input DataFrame.
    group_col (str): The name of the column containing group labels.
    value_col (str): The name of the column containing values to plot.

    Returns:
    plt.Figure: The matplotlib figure object.
    """
    from statsmodels.formula.api import ols

    sanitized_value_col = value_col.replace('-', '_').replace(' ', '_')
    sanitized_group_col = group_col.replace('-', '_').replace(' ', '_')
    
    data = data.dropna(subset=[value_col, group_col])
    data = data[np.isfinite(data[value_col])]
    
    if data[sanitized_group_col].nunique() < 2 or len(data) < 2:
        print(f"Not enough data to fit OLS model for {value_col}. Skipping.")
        return None
    
    data = data.rename(columns={value_col: sanitized_value_col, group_col: sanitized_group_col})
    
    formula = f'{sanitized_value_col} ~ {sanitized_group_col}'
    try:
        model = ols(formula, data=data).fit()
    except ValueError as e:
        print(f"Error fitting OLS model for {value_col}: {e}")
        return None
    
    residuals = model.resid
    fitted = model.fittedvalues
    
    fig, axs = plt.subplots(2, 2, figsize=(12, 10))
    
    sns.histplot(residuals, bins=15, kde=True, ax=axs[0, 0])
    axs[0, 0].set_title('Histogram of Residuals')
    
    stats.probplot(residuals, dist="norm", plot=axs[0, 1])
    axs[0, 1].set_title('Normal Probability Plot')
    
    sns.scatterplot(x=fitted, y=residuals, ax=axs[1, 0])
    axs[1, 0].axhline(0, color='r', linestyle='--')
    axs[1, 0].set_title('Residuals vs Fitted Values')
    
    sns.lineplot(x=np.arange(len(residuals)), y=residuals, ax=axs[1, 1])
    axs[1, 1].set_title('Residuals vs Order')
    axs[1, 1].axhline(0, color='r', linestyle='--')
    
    plt.tight_layout()
    fig=plt.gcf()
    plt.close(fig)
    return fig

# Adjusted function to plot all visualizations for each numerical column against 'GROUP' for each Attribute
def plot_all_visualizations(data):
    """
    Plots all relevant visualizations for the numeric columns in the data.

    Parameters:
    data (pd.DataFrame): The input DataFrame.
    """
    group_col = 'GROUP'
    numeric_columns = data.select_dtypes(include='number').columns
    
    # Loop through each attribute in the data
    for attr in data['Attributes'].unique():
        attr_data = data[data['Attributes'] == attr]
        
        for value_col in numeric_columns:
            if value_col != group_col:
                print(f"Plotting for Attribute: {attr}, Column: {value_col}")
                
                # Create Tukey's HSD plot
                fig = plot_tukey_CIs(attr_data, group_col, value_col, attr)
                if fig:
                    print(f"Storing Tukey's HSD plot for {attr}_{value_col} under 'tukey_CIs'")
                    plots_dict.setdefault('tukey_CIs', []).append((f'{attr}_{value_col}', fig))
                
                # Create Interval plot
                fig = plot_interval(attr_data, group_col, value_col, attr)
                if fig:
                    print(f"Storing Interval plot for {attr}_{value_col} under 'interval'")
                    plots_dict.setdefault('interval', []).append((f'{attr}_{value_col}', fig))
                
                # Create Individual Value plot
                fig = plot_individual_value(attr_data, group_col, value_col, attr)
                if fig:
                    print(f"Storing Individual Value plot for {attr}_{value_col} under 'individual_value'")
                    plots_dict.setdefault('individual_value', []).append((f'{attr}_{value_col}', fig))
                
                # Create Boxplot
                fig = plot_boxplot(attr_data, group_col, value_col, attr)
                if fig:
                    print(f"Storing Boxplot for {attr}_{value_col} under 'boxplot'")
                    plots_dict.setdefault('boxplot', []).append((f'{attr}_{value_col}', fig))
                
                # Create Residuals plot
                fig = plot_residuals(attr_data, group_col, value_col, attr)
                if fig:
                    print(f"Storing Residuals plot for {attr}_{value_col} under 'residuals'")
                    plots_dict.setdefault('residuals', []).append((f'{attr}_{value_col}', fig))

                    
def main():
    """
    The main function that runs the Streamlit app, handles file uploads, and performs data analysis.
    """
    st.image("Carlisle_MasterLogo_RGB.jpg", width=500)
    
    st.markdown("""# Data Analysis Automation&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<a href="https://dummy-link.com" style="font-size:20px; color:#007BFF; font-style:italic;">SOP &#9432;</a>""", unsafe_allow_html=True)

    uploaded_file = st.file_uploader("**Upload your Excel file**", type=["xlsx"])
    
    if uploaded_file:
        data = pd.read_excel(uploaded_file, sheet_name='Data set')
        
        if 'DESCRIPTION' not in data.columns:
            st.error("The column 'DESCRIPTION' does not exist in the dataset.")
        else:
            global final_df_global, new_df, table1_df_global, outliers_df_global, granularity_choice, plots_dict

            st.session_state.updated_data = data.copy()
            final_df_global = None
            new_df = None
            table1_df_global = None
            outliers_df_global = pd.DataFrame(columns=['Attributes', 'Column', 'Outlier'])
            granularity_choice = 'Upto Top and Bottom Facer'
            plots_dict = {}

            exceptions = ['DESCRIPTION', 'SURFACTANT', 'SAP']
            data = convert_columns_to_numeric(data, exceptions=exceptions)

            if 'SAP-Desc' not in st.session_state.updated_data.columns:
                st.session_state.updated_data['SAP-Desc'] = st.session_state.updated_data['DESCRIPTION']
            if 'Merge-Key' not in st.session_state.updated_data.columns:
                st.session_state.updated_data['Merge-Key'] = np.nan

            granularity_choice = st.selectbox('**Select Granularity**', ['Select Granularity', 'Upto Top and Bottom Facer', 'Upto Dimensions'])

            if granularity_choice == 'Upto Top and Bottom Facer':
                description_options = data['DESCRIPTION'].unique().tolist()
                description_options.insert(0, 'Select Descriptions')
                selected_descriptions = st.multiselect('**Select Descriptions**', description_options)
                
                if 'Select Descriptions' not in selected_descriptions and selected_descriptions:
                    filtered_df = filter_dataframe(selected_descriptions)
                    
                    column_options = create_column_dropdown(filtered_df)
                    column_options.insert(0, 'Select Columns')
                    selected_columns = st.multiselect('**Select Columns**', column_options)
                    
                    if 'Select Columns' not in selected_columns and selected_columns:
                        final_df = select_columns(filtered_df, selected_columns)
                        
                        target_column = st.selectbox('**Select Target Column**', ['Select Target Column'] + final_df.columns.tolist(), index=0)
                        
                        if target_column != 'Select Target Column':
                            new_df = set_target_column(final_df, target_column)
                            detect_outliers(new_df, granularity_choice)
                            final_df_global = prompt_outlier_removal(new_df, outliers_df_global, granularity_choice)
                            st.write("**Final DataFrame after Removing Outliers:**")
                            st.write(final_df_global)
                            calculate_and_display_statistics(final_df_global, granularity_choice)
                            perform_anderson_darling_test(final_df_global)
                            plot_all_visualizations(final_df_global)
                            
                            if table1_df_global is not None:
                                unique_groups = table1_df_global['GROUP'].unique().tolist()
                                unique_groups.insert(0, 'Select')
                                trial_group = st.selectbox("**Please select the Trial group from the above unique GROUP values:**", unique_groups, index=0)
                                incumbent_group = st.selectbox("**Please select the Incumbent group from the above unique GROUP values:**", unique_groups, index=0)
                                
                                if trial_group != 'Select' and incumbent_group != 'Select':
                                    display_results(user_choice_1, trial_group, incumbent_group)
            
            elif granularity_choice == 'Upto Dimensions':
                interactive_sap_filter_and_select()
                
                if final_df_global is not None:
                    final_df_global = prompt_outlier_removal(final_df_global, outliers_df_global, granularity_choice)
                    st.write("**Final DataFrame after Removing Outliers:**")
                    st.write(final_df_global)
                    perform_anderson_darling_test(final_df_global)
                    plot_all_visualizations(final_df_global)
                    
                    if table1_df_global is not None:
                        unique_groups = table1_df_global['GROUP'].unique().tolist()
                        unique_groups.insert(0, 'Select')
                        trial_group = st.selectbox("Please select the Trial group from the above unique GROUP values:", unique_groups, index=0)
                        incumbent_group = st.selectbox("Please select the Incumbent group from the above unique GROUP values:", unique_groups, index=0)
                        if trial_group != 'Select' and incumbent_group != 'Select':
                            display_results(user_choice_1, trial_group, incumbent_group)

if __name__ == "__main__":
    main()